from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models import Movimentacao
from app.services import dashboard as dash
from app.utils.formatters import formatar_data, formatar_moeda, nome_mes


def movimentacoes_periodo(usuario_id: int, inicio: date, fim: date):
    return (
        Movimentacao.query.filter(
            Movimentacao.usuario_id == usuario_id,
            Movimentacao.ativo.is_(True),
            Movimentacao.data >= inicio,
            Movimentacao.data <= fim,
        )
        .order_by(Movimentacao.data.asc(), Movimentacao.id.asc())
        .all()
    )


def montar(usuario_id: int, inicio: date, fim: date) -> dict:
    resumo = dash.resumo_periodo(usuario_id, inicio, fim)
    categorias = dash.gastos_por_categoria(usuario_id, inicio, fim)
    lancamentos = movimentacoes_periodo(usuario_id, inicio, fim)
    linhas_cat = list(
        zip(categorias.get("labels") or [], categorias.get("valores") or [], categorias.get("cores") or [])
    )
    titulo_periodo = (
        nome_mes(inicio.year, inicio.month)
        if inicio.month == fim.month and inicio.year == fim.year and inicio.day == 1
        else f"{formatar_data(inicio)} a {formatar_data(fim)}"
    )
    return {
        "resumo": resumo,
        "categorias": linhas_cat,
        "lancamentos": lancamentos,
        "inicio": inicio,
        "fim": fim,
        "titulo_periodo": titulo_periodo,
    }


def _estilos():
    estilos = getSampleStyleSheet()
    estilos.add(
        ParagraphStyle(
            name="TituloFin",
            fontName="Helvetica-Bold",
            fontSize=16,
            textColor=colors.HexColor("#0f766e"),
            spaceAfter=6,
        )
    )
    estilos.add(
        ParagraphStyle(
            name="SubFin",
            fontName="Helvetica",
            fontSize=10,
            textColor=colors.HexColor("#64748b"),
            spaceAfter=12,
        )
    )
    return estilos


def gerar_pdf(usuario_id: int, inicio: date, fim: date, nome_usuario: str) -> BytesIO:
    dados = montar(usuario_id, inicio, fim)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
        title="Relatório FinUP",
    )
    estilos = _estilos()
    story = [
        Paragraph("FinUP — Relatório financeiro", estilos["TituloFin"]),
        Paragraph(
            f"{nome_usuario} · {dados['titulo_periodo']} · {formatar_data(inicio)} a {formatar_data(fim)}",
            estilos["SubFin"],
        ),
    ]
    resumo = dados["resumo"]
    kpi = [
        ["Receitas", "Despesas", "Investimentos", "Saldo do período"],
        [
            formatar_moeda(resumo["receitas"]),
            formatar_moeda(resumo["despesas"]),
            formatar_moeda(resumo["investimentos"]),
            formatar_moeda(resumo["saldo_periodo"]),
        ],
    ]
    tabela_kpi = Table(kpi, colWidths=[4 * cm, 4 * cm, 4 * cm, 4.2 * cm])
    tabela_kpi.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f1f5f9")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.extend([tabela_kpi, Spacer(1, 14)])

    story.append(Paragraph("Gastos por categoria", estilos["Heading2"]))
    cat_rows = [["Categoria", "Valor"]]
    if dados["categorias"]:
        cat_rows.extend([[nome, formatar_moeda(valor)] for nome, valor, _ in dados["categorias"]])
    else:
        cat_rows.append(["Sem gastos categorizados", "—"])
    tabela_cat = Table(cat_rows, colWidths=[12 * cm, 4.2 * cm])
    tabela_cat.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([tabela_cat, Spacer(1, 14)])

    story.append(Paragraph("Lançamentos", estilos["Heading2"]))
    mov_rows = [["Data", "Descrição", "Tipo", "Valor"]]
    for mov in dados["lancamentos"][:250]:
        mov_rows.append(
            [
                formatar_data(mov.data),
                (mov.descricao or "")[:42],
                mov.tipo_label,
                formatar_moeda(mov.valor),
            ]
        )
    if len(mov_rows) == 1:
        mov_rows.append(["—", "Nenhum lançamento no período", "—", "—"])
    tabela_mov = Table(mov_rows, colWidths=[2.6 * cm, 8.2 * cm, 2.6 * cm, 2.8 * cm])
    tabela_mov.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tabela_mov)
    doc.build(story)
    buffer.seek(0)
    return buffer


def gerar_excel(usuario_id: int, inicio: date, fim: date) -> BytesIO:
    dados = montar(usuario_id, inicio, fim)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "FinUP — Relatório"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{formatar_data(inicio)} a {formatar_data(fim)}"
    ws["A4"] = "Receitas"
    ws["B4"] = formatar_moeda(dados["resumo"]["receitas"])
    ws["A5"] = "Despesas"
    ws["B5"] = formatar_moeda(dados["resumo"]["despesas"])
    ws["A6"] = "Investimentos"
    ws["B6"] = formatar_moeda(dados["resumo"]["investimentos"])
    ws["A7"] = "Saldo do período"
    ws["B7"] = formatar_moeda(dados["resumo"]["saldo_periodo"])
    ws["A8"] = "Saldo nas contas"
    ws["B8"] = formatar_moeda(dados["resumo"]["saldo_contas"])

    ws_cat = wb.create_sheet("Categorias")
    ws_cat.append(["Categoria", "Valor"])
    for nome, valor, _ in dados["categorias"]:
        ws_cat.append([nome, float(valor)])

    ws_mov = wb.create_sheet("Lancamentos")
    ws_mov.append(["Data", "Descricao", "Tipo", "Categoria", "Conta", "Forma", "Valor"])
    for mov in dados["lancamentos"]:
        ws_mov.append(
            [
                formatar_data(mov.data),
                mov.descricao,
                mov.tipo_label,
                mov.categoria.nome if mov.categoria else "",
                mov.conta.nome if mov.conta else "",
                mov.forma_label,
                float(mov.valor),
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
